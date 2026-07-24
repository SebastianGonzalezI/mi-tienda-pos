import os
from openpyxl import load_workbook


def cargar_productos_desde_excel(ruta=None):
    if ruta is None:
        ruta = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'productos_ejemplo.xlsx')

    if not os.path.exists(ruta):
        return []

    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    productos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        producto = {
            'codigo': str(row[0]) if row[0] else '',
            'nombre': str(row[1]) if row[1] else '',
            'cantidad': int(row[2]) if row[2] else 0,
            'precio_sin_iva': float(row[3]) if row[3] else 0,
            'impuesto': float(row[4]) if row[4] else 12,
            'precio_con_iva': float(row[5]) if row[5] else 0,
            'descuento': float(row[6]) if row[6] else 0,
        }
        productos.append(producto)

    return productos


def crear_excel_ejemplo(ruta=None):
    if ruta is None:
        ruta = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'productos_ejemplo.xlsx')

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ['Código', 'Nombre', 'Cantidad', 'Precio Sin IVA', 'IVA (%)', 'Precio Con IVA', 'Descuento']
    ws.append(headers)

    productos = [
        ('VET001', 'Vacuna Triple Felina', 50, 25.00, 12, 28.00, 0),
        ('VET002', 'Desparasitante Oral Perros', 80, 15.00, 12, 16.80, 0),
        ('VET003', 'Shampoo Antipulgas', 40, 18.50, 12, 20.72, 0),
        ('VET004', 'Collar Antipulgas', 30, 22.00, 12, 24.64, 0),
        ('VET005', 'Comida Premium Perro 3kg', 25, 35.00, 12, 39.20, 0),
        ('VET006', 'Comida Gato Adulto 1.5kg', 20, 28.50, 12, 31.92, 0),
        ('VET007', 'Jeringas 5ml (x100)', 60, 12.00, 12, 13.44, 0),
        ('VET008', 'Vendas Elásticas (x10)', 45, 8.50, 12, 9.52, 0),
        ('VET009', 'Antibiótico Amoxicilina 500mg', 35, 42.00, 12, 47.04, 0),
        ('VET010', 'Guantes Quirúrgicos (x50)', 100, 6.00, 12, 6.72, 0),
    ]

    for p in productos:
        ws.append(p)

    wb.save(ruta)
    return ruta
